"""
Task History Database for Gastos Robot GUI

SQLite database for storing and retrieving task execution history.
Supports search, filtering, statistics, and export functionality.
"""

import sqlite3
import json
import csv
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
import threading


class TaskHistoryDB:
    """SQLite database for task history."""

    def __init__(self, db_path: str = "gastos_task_history.db"):
        """Initialize the database."""
        self.db_path = db_path
        self.lock = threading.Lock()
        self._init_database()

    def _init_database(self):
        """Create the database tables if they don't exist."""
        with self.lock:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cursor.execute("""
                CREATE TABLE IF NOT EXISTS task_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    task_id TEXT NOT NULL,
                    operation_type TEXT,
                    operation_number TEXT,
                    date TEXT,
                    cash_register TEXT,
                    third_party TEXT,
                    nature TEXT,
                    amount REAL,
                    description TEXT,
                    total_line_items INTEGER,
                    status TEXT NOT NULL,
                    started_at TEXT NOT NULL,
                    completed_at TEXT,
                    duration_seconds REAL,
                    error_message TEXT,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # Create indexes for faster searches
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_task_id ON task_history(task_id)
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_operation_number ON task_history(operation_number)
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_status ON task_history(status)
            """)
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_created_at ON task_history(created_at)
            """)

            conn.commit()
            conn.close()

    def add_task(self, task_data: Dict[str, Any]) -> bool:
        """
        Add a completed task to history.

        Args:
            task_data: Dictionary containing task information

        Returns:
            bool: True if successful, False otherwise
        """
        try:
            with self.lock:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()

                cursor.execute("""
                    INSERT INTO task_history (
                        task_id, operation_type, operation_number, date, cash_register,
                        third_party, nature, amount, description, total_line_items,
                        status, started_at, completed_at, duration_seconds, error_message
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    task_data.get('task_id'),
                    task_data.get('operation_type'),
                    task_data.get('operation_number'),
                    task_data.get('date'),
                    task_data.get('cash_register'),
                    task_data.get('third_party'),
                    task_data.get('nature'),
                    task_data.get('amount'),
                    task_data.get('description'),
                    task_data.get('total_line_items'),
                    task_data.get('status'),
                    task_data.get('started_at'),
                    task_data.get('completed_at'),
                    task_data.get('duration_seconds'),
                    task_data.get('error_message')
                ))

                conn.commit()
                conn.close()
                return True

        except Exception as e:
            print(f"Error adding task to history: {e}")
            return False

    def get_all_tasks(self, limit: int = 100, status_filter: Optional[str] = None) -> List[Dict[str, Any]]:
        """
        Get all tasks from history, most recent first.

        Args:
            limit: Maximum number of tasks to return
            status_filter: Optional status filter ('completed', 'failed', etc.)

        Returns:
            List of task dictionaries
        """
        try:
            with self.lock:
                conn = sqlite3.connect(self.db_path)
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()

                if status_filter:
                    cursor.execute("""
                        SELECT * FROM task_history
                        WHERE status = ?
                        ORDER BY created_at DESC
                        LIMIT ?
                    """, (status_filter, limit))
                else:
                    cursor.execute("""
                        SELECT * FROM task_history
                        ORDER BY created_at DESC
                        LIMIT ?
                    """, (limit,))

                rows = cursor.fetchall()
                tasks = [dict(row) for row in rows]

                conn.close()
                return tasks

        except Exception as e:
            print(f"Error getting tasks: {e}")
            return []

    def search_tasks(self, search_term: str, limit: int = 100) -> List[Dict[str, Any]]:
        """
        Search tasks by task_id, operation_number, or third_party.

        Args:
            search_term: Search string
            limit: Maximum number of results

        Returns:
            List of matching task dictionaries
        """
        try:
            with self.lock:
                conn = sqlite3.connect(self.db_path)
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()

                search_pattern = f"%{search_term}%"
                cursor.execute("""
                    SELECT * FROM task_history
                    WHERE task_id LIKE ?
                       OR operation_number LIKE ?
                       OR third_party LIKE ?
                    ORDER BY created_at DESC
                    LIMIT ?
                """, (search_pattern, search_pattern, search_pattern, limit))

                rows = cursor.fetchall()
                tasks = [dict(row) for row in rows]

                conn.close()
                return tasks

        except Exception as e:
            print(f"Error searching tasks: {e}")
            return []

    def get_statistics(self) -> Dict[str, Any]:
        """
        Get overall statistics from task history.

        Returns:
            Dictionary with statistics
        """
        try:
            with self.lock:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()

                # Total tasks
                cursor.execute("SELECT COUNT(*) FROM task_history")
                total_tasks = cursor.fetchone()[0]

                # Completed tasks
                cursor.execute("SELECT COUNT(*) FROM task_history WHERE status = 'completed'")
                completed = cursor.fetchone()[0]

                # Failed tasks
                cursor.execute("SELECT COUNT(*) FROM task_history WHERE status IN ('failed', 'error')")
                failed = cursor.fetchone()[0]

                # Average duration
                cursor.execute("SELECT AVG(duration_seconds) FROM task_history WHERE duration_seconds IS NOT NULL")
                avg_duration = cursor.fetchone()[0] or 0

                conn.close()

                return {
                    'total_tasks': total_tasks,
                    'completed': completed,
                    'failed': failed,
                    'avg_duration': avg_duration
                }

        except Exception as e:
            print(f"Error getting statistics: {e}")
            return {
                'total_tasks': 0,
                'completed': 0,
                'failed': 0,
                'avg_duration': 0
            }

    def export_to_json(self, filepath: str, limit: int = 1000) -> bool:
        """Export task history to JSON file."""
        try:
            tasks = self.get_all_tasks(limit=limit)

            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(tasks, f, indent=2, ensure_ascii=False)

            return True

        except Exception as e:
            print(f"Error exporting to JSON: {e}")
            return False

    def export_to_csv(self, filepath: str, limit: int = 1000) -> bool:
        """Export task history to CSV file."""
        try:
            tasks = self.get_all_tasks(limit=limit)

            if not tasks:
                return False

            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=tasks[0].keys())
                writer.writeheader()
                writer.writerows(tasks)

            return True

        except Exception as e:
            print(f"Error exporting to CSV: {e}")
            return False

    def export_to_excel(self, filepath: str, limit: int = 1000) -> bool:
        """Export task history to Excel file."""
        try:
            # Try to import openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                print("openpyxl not installed. Install with: pip install openpyxl")
                return False

            tasks = self.get_all_tasks(limit=limit)

            if not tasks:
                return False

            wb = Workbook()
            ws = wb.active
            ws.title = "Gastos Task History"

            # Headers
            headers = list(tasks[0].keys())
            ws.append(headers)

            # Style headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Add data
            for task in tasks:
                ws.append(list(task.values()))

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(filepath)
            return True

        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return False

    def clear_history(self) -> bool:
        """Clear all task history."""
        try:
            with self.lock:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                cursor.execute("DELETE FROM task_history")
                conn.commit()
                conn.close()
                return True

        except Exception as e:
            print(f"Error clearing history: {e}")
            return False


# Global instance
_db_instance: Optional[TaskHistoryDB] = None
_db_lock = threading.Lock()


def get_task_history_db() -> TaskHistoryDB:
    """Get the global TaskHistoryDB instance (singleton pattern)."""
    global _db_instance
    if _db_instance is None:
        with _db_lock:
            if _db_instance is None:
                _db_instance = TaskHistoryDB()
    return _db_instance
